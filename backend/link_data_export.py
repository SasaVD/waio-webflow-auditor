"""
Link Data Export — Screaming Frog-style Excel and CSV exports
=============================================================

Generates multi-sheet Excel workbooks and CSV ZIP archives from
link graph data and TIPR analysis results.
"""

import csv
import io
import zipfile
import logging
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def _short_url(url: str) -> str:
    """Return a display-friendly URL path."""
    for prefix in ("https://", "http://"):
        if url.startswith(prefix):
            url = url[len(prefix):]
    return url.rstrip("/") or "/"


def _build_pages_rows(graph: dict, tipr: dict | None) -> list[list]:
    """Sheet 1: Pages (Node List)."""
    header = [
        "URL", "Title", "Status Code", "Crawl Depth", "Inbound Links",
        "Unique Inbound", "Outbound Links", "PR Score (0-100)",
        "CheiRank Score (0-100)", "TIPR Rank", "TIPR Classification",
        "Cluster", "Is Orphan", "Is Hub (>10 inbound)",
    ]
    rows = [header]

    nodes = graph.get("nodes") or []
    tipr_lookup: dict[str, dict] = {}
    if tipr and tipr.get("pages"):
        for p in tipr["pages"]:
            tipr_lookup[p["url"]] = p

    for node in nodes:
        url = node.get("id", "")
        tp = tipr_lookup.get(url, {})
        inbound = node.get("inbound", 0)
        rows.append([
            url,
            node.get("label", ""),
            node.get("status_code", ""),
            node.get("depth", ""),
            inbound,
            inbound,  # unique inbound = same as inbound for internal links
            node.get("outbound", 0),
            round(tp.get("pagerank_score", 0), 1) if tp else "",
            round(tp.get("cheirank_score", 0), 1) if tp else "",
            tp.get("tipr_rank", "") if tp else "",
            tp.get("classification", "") if tp else "",
            tp.get("cluster", node.get("cluster", "")),
            "Yes" if inbound == 0 else "No",
            "Yes" if inbound > 10 else "No",
        ])
    return rows


def _build_links_rows(graph: dict) -> list[list]:
    """Sheet 2: Links (Edge List)."""
    header = ["Source URL", "Target URL", "Anchor Text", "Link Type", "Status Code of Target"]
    rows = [header]
    for link in (graph.get("links") or []):
        src = link.get("source") if isinstance(link.get("source"), str) else (link.get("source") or {}).get("id", "")
        tgt = link.get("target") if isinstance(link.get("target"), str) else (link.get("target") or {}).get("id", "")
        rows.append([
            src, tgt,
            link.get("anchor", ""),
            link.get("type", "internal"),
            link.get("target_status_code", ""),
        ])
    return rows


def _build_tipr_rows(tipr: dict) -> list[list]:
    """Sheet 3: TIPR Analysis."""
    header = [
        "URL", "PageRank", "PageRank Score", "CheiRank", "CheiRank Score",
        "TIPR Rank", "Classification", "Inbound", "Outbound", "Delta (In-Out)", "Cluster",
    ]
    rows = [header]
    for p in (tipr.get("pages") or []):
        rows.append([
            p["url"],
            p.get("pagerank", 0),
            round(p.get("pagerank_score", 0), 1),
            p.get("cheirank", 0),
            round(p.get("cheirank_score", 0), 1),
            p.get("tipr_rank", 0),
            p.get("classification", ""),
            p.get("inbound_count", 0),
            p.get("outbound_count", 0),
            p.get("inbound_count", 0) - p.get("outbound_count", 0),
            p.get("cluster", ""),
        ])
    return rows


def _build_recommendations_rows(tipr: dict) -> list[list]:
    """Sheet 4: Recommendations."""
    header = [
        "Priority", "Group", "Type", "Source URL", "Target URL",
        "Reason", "Expected Impact", "Source PR", "Target PR",
        "Source Outlinks", "Content Relevance",
    ]
    rows = [header]
    for r in (tipr.get("recommendations") or []):
        rows.append([
            r.get("priority", ""),
            r.get("group", ""),
            r.get("type", ""),
            r.get("source_url", ""),
            r.get("target_url", ""),
            r.get("reason", ""),
            r.get("expected_impact", ""),
            r.get("source_pr_score", 0),
            r.get("target_pr_score", 0),
            r.get("source_outlinks", 0),
            r.get("content_relevance", 0),
        ])
    return rows


def _build_orphans_rows(graph: dict, tipr: dict | None) -> list[list]:
    """Sheet 5: Orphan Pages."""
    header = ["URL", "Cluster", "Suggested Source 1", "Suggested Source 2", "Suggested Source 3"]
    rows = [header]

    nodes = graph.get("nodes") or []
    orphan_urls = [n["id"] for n in nodes if n.get("inbound", 0) == 0]

    # Build recommendation targets → sources map
    rec_map: dict[str, list[str]] = {}
    if tipr and tipr.get("recommendations"):
        for r in tipr["recommendations"]:
            if r.get("type") == "add_link" and r.get("target_url"):
                rec_map.setdefault(r["target_url"], []).append(r.get("source_url", ""))

    tipr_lookup: dict[str, dict] = {}
    if tipr and tipr.get("pages"):
        for p in tipr["pages"]:
            tipr_lookup[p["url"]] = p

    for url in orphan_urls:
        tp = tipr_lookup.get(url, {})
        suggested = rec_map.get(url, [])[:3]
        while len(suggested) < 3:
            suggested.append("")
        rows.append([url, tp.get("cluster", ""), *suggested])
    return rows


def _build_summary_rows(graph: dict, tipr: dict | None) -> list[list]:
    """Sheet 6: Summary stats."""
    nodes = graph.get("nodes") or []
    links = graph.get("links") or []
    stats = graph.get("stats") or {}

    total_pages = len(nodes)
    total_links = len(links)
    orphan_count = sum(1 for n in nodes if n.get("inbound", 0) == 0)
    hub_count = sum(1 for n in nodes if n.get("inbound", 0) > 10)
    max_depth = stats.get("max_depth", 0)
    avg_links = round(total_links / total_pages, 1) if total_pages > 0 else 0

    rows = [["Metric", "Value"]]
    rows.append(["Total Pages", total_pages])
    rows.append(["Total Links", total_links])
    rows.append(["Orphan Pages", orphan_count])
    rows.append(["Hub Pages (>10 inbound)", hub_count])

    if tipr and tipr.get("summary"):
        s = tipr["summary"]
        rows.append(["Stars", s.get("stars", 0)])
        rows.append(["Hoarders", s.get("hoarders", 0)])
        rows.append(["Wasters", s.get("wasters", 0)])
        rows.append(["Dead Weight", s.get("dead_weight", 0)])

    rows.append(["Max Crawl Depth", max_depth])
    rows.append(["Avg Links/Page", avg_links])

    if tipr and tipr.get("recommendations"):
        rows.append(["Total Recommendations", len(tipr["recommendations"])])

    return rows


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_link_data_excel(report: dict[str, Any]) -> bytes:
    """Generate a multi-sheet Excel workbook with link intelligence data."""
    graph = (report.get("link_analysis") or {}).get("graph") or {}
    tipr = report.get("tipr_analysis")

    wb = Workbook()

    def _add_sheet(name: str, data: list[list]):
        ws = wb.create_sheet(title=name)
        for row in data:
            ws.append(row)
        # Auto-width (approximate)
        for col_idx, _ in enumerate(data[0] if data else [], 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row in data[:50]:
                if col_idx <= len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1])))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # Remove default sheet
    wb.remove(wb.active)

    _add_sheet("Pages (Node List)", _build_pages_rows(graph, tipr))
    _add_sheet("Links (Edge List)", _build_links_rows(graph))
    if tipr:
        _add_sheet("TIPR Analysis", _build_tipr_rows(tipr))
        _add_sheet("Recommendations", _build_recommendations_rows(tipr))
    _add_sheet("Orphan Pages", _build_orphans_rows(graph, tipr))
    _add_sheet("Summary", _build_summary_rows(graph, tipr))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_link_data_csv_zip(report: dict[str, Any]) -> bytes:
    """Generate a ZIP archive containing CSV files for link data."""
    graph = (report.get("link_analysis") or {}).get("graph") or {}
    tipr = report.get("tipr_analysis")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        def _write_csv(filename: str, data: list[list]):
            csv_buf = io.StringIO()
            writer = csv.writer(csv_buf)
            writer.writerows(data)
            zf.writestr(filename, csv_buf.getvalue())

        _write_csv("pages-node-list.csv", _build_pages_rows(graph, tipr))
        _write_csv("links-edge-list.csv", _build_links_rows(graph))
        if tipr:
            _write_csv("tipr-analysis.csv", _build_tipr_rows(tipr))
            _write_csv("recommendations.csv", _build_recommendations_rows(tipr))
        _write_csv("orphan-pages.csv", _build_orphans_rows(graph, tipr))
        _write_csv("summary.csv", _build_summary_rows(graph, tipr))

    return buf.getvalue()
